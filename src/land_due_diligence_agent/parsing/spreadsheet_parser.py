"""Spreadsheet and CSV extraction helpers."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook


def extract_xlsx_text(path: Path) -> tuple[str, dict[str, int], list[str], list[dict[str, str | int | None]]]:
    """Extract cell text from each worksheet in an XLSX workbook."""

    warnings: list[str] = []
    sections: list[str] = []
    chunk_records: list[dict[str, str | int | None]] = []
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    sheet_count = len(workbook.sheetnames)

    try:
        for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
            rows: list[str] = []
            for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                values = [_stringify_cell(value) for value in row if _stringify_cell(value)]
                if values:
                    row_text = " | ".join(values)
                    rows.append(row_text)
                    chunk_records.append(
                        {
                            "chunk_id": f"sheet-{sheet_index:02d}-row-{row_index:04d}",
                            "page_number": None,
                            "text": f"Sheet: {worksheet.title}\n{row_text}",
                        }
                    )
            if rows:
                sections.append(f"Sheet: {worksheet.title}\n" + "\n".join(rows))
    finally:
        workbook.close()

    if not sections:
        warnings.append("No XLSX content extracted.")

    return "\n\n".join(sections), {"sheet_count": sheet_count}, warnings, chunk_records


def extract_xls_text(path: Path) -> tuple[str, dict[str, int], list[str], list[dict[str, str | int | None]]]:
    """Extract cell text from each worksheet in an XLS workbook."""

    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - dependency issue
        raise RuntimeError(
            "XLS parsing requires xlrd>=2.0.1. Install project dependencies with pip install -r requirements.txt."
        ) from exc

    warnings: list[str] = []
    sections: list[str] = []
    chunk_records: list[dict[str, str | int | None]] = []
    workbook = xlrd.open_workbook(str(path), on_demand=True)
    sheet_count = workbook.nsheets

    try:
        for sheet_index in range(workbook.nsheets):
            worksheet = workbook.sheet_by_index(sheet_index)
            rows: list[str] = []
            for row_index in range(worksheet.nrows):
                values = [
                    _stringify_cell(worksheet.cell_value(row_index, column_index))
                    for column_index in range(worksheet.ncols)
                ]
                values = [value for value in values if value]
                if values:
                    row_text = " | ".join(values)
                    rows.append(row_text)
                    chunk_records.append(
                        {
                            "chunk_id": f"sheet-{sheet_index + 1:02d}-row-{row_index + 1:04d}",
                            "page_number": None,
                            "text": f"Sheet: {worksheet.name}\n{row_text}",
                        }
                    )
            if rows:
                sections.append(f"Sheet: {worksheet.name}\n" + "\n".join(rows))
    finally:
        release_resources = getattr(workbook, "release_resources", None)
        if callable(release_resources):
            release_resources()

    if not sections:
        warnings.append("No XLS content extracted.")

    return "\n\n".join(sections), {"sheet_count": sheet_count}, warnings, chunk_records


def extract_csv_text(path: Path) -> tuple[str, dict[str, int], list[str], list[dict[str, str | int | None]]]:
    """Extract row text from a CSV file."""

    warnings: list[str] = []
    rows: list[str] = []
    chunk_records: list[dict[str, str | int | None]] = []
    encoding_used = "utf-8-sig"

    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.reader(handle)
                for row_index, row in enumerate(reader, start=1):
                    values = [value.strip() for value in row if value.strip()]
                    if values:
                        row_text = " | ".join(values)
                        rows.append(row_text)
                        chunk_records.append(
                            {
                                "chunk_id": f"row-{row_index:04d}",
                                "page_number": None,
                                "text": row_text,
                            }
                        )
            encoding_used = encoding
            break
        except UnicodeDecodeError:
            rows.clear()
            chunk_records.clear()
            continue

    if not rows:
        warnings.append("No CSV content extracted.")

    return "\n".join(rows), {"row_count": len(rows), "encoding": encoding_used}, warnings, chunk_records


def _stringify_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
